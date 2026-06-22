#!/usr/bin/env python3
"""值班日历 Web 应用 — Flask + Bootstrap 5"""

import json
import os
from datetime import date, timedelta
from io import BytesIO

from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from chinese_calendar import is_holiday, get_holiday_detail

app = Flask(__name__)

CONFIG_PATH = os.environ.get("CONFIG_PATH", "config.json")

HOLIDAY_NAME_MAP = {
    "New Year's Day": "元旦",
    "Spring Festival": "春节",
    "Tomb-sweeping Day": "清明节",
    "Labour Day": "劳动节",
    "Dragon Boat Festival": "端午节",
    "National Day": "国庆节",
    "Mid-autumn Festival": "中秋节",
    "Anti-Fascist 70th Day": "抗战胜利纪念日",
    "节假日": "节假日",
}


# ── 工具函数 ──────────────────────────────────────────────

def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def get_duty_class(cfg, target_date):
    """根据起始日期和班数计算某天是哪个班值班"""
    start = date.fromisoformat(cfg["start_date"])
    classes = cfg["classes"]
    n = len(classes)
    if n == 0:
        return None
    delta = (target_date - start).days
    if delta < 0:
        # 起始日期之前的日期，反向推算
        idx = (delta % n)
    else:
        idx = delta % n
    return classes[idx]


def get_holiday_info(target_date):
    """获取节假日信息，返回 (是否节假日, 节假日名/None)"""
    try:
        if is_holiday(target_date):
            _, name = get_holiday_detail(target_date)
            name = name or "节假日"
            return True, HOLIDAY_NAME_MAP.get(name, name)
        else:
            return False, None
    except Exception:
        return False, None


def is_makeup_day(target_date):
    """判断是否是补班日（周末但需上班）"""
    try:
        # 周末但 is_workday=True 且不是节假日 = 补班日
        if target_date.weekday() >= 5:
            from chinese_calendar import is_workday
            if is_workday(target_date) and not is_holiday(target_date):
                return True
        return False
    except Exception:
        return False


def build_month_data(cfg, year, month):
    """构建某月的日历数据"""
    first_day = date(year, month, 1)
    if month == 12:
        next_month_first = date(year + 1, 1, 1)
    else:
        next_month_first = date(year, month + 1, 1)
    days_in_month = (next_month_first - first_day).days

    # weekday of first day (0=Mon)
    start_weekday = first_day.weekday()

    days = []
    for d in range(1, days_in_month + 1):
        dt = date(year, month, d)
        duty = get_duty_class(cfg, dt)
        is_hol, hol_name = get_holiday_info(dt)
        makeup = is_makeup_day(dt)
        days.append({
            "date": dt.isoformat(),
            "day": d,
            "weekday": dt.weekday(),
            "duty_class": duty,
            "is_holiday": is_hol,
            "holiday_name": hol_name,
            "is_makeup": makeup,
        })

    return {
        "year": year,
        "month": month,
        "start_weekday": start_weekday,
        "days": days,
    }


# ── 路由 ──────────────────────────────────────────────────

@app.route("/")
def index():
    cfg = load_config()
    today = date.today()
    # 显示当前月和下月
    m1_year, m1_month = today.year, today.month
    if m1_month == 12:
        m2_year, m2_month = m1_year + 1, 1
    else:
        m2_year, m2_month = m1_year, m1_month + 1

    month1 = build_month_data(cfg, m1_year, m1_month)
    month2 = build_month_data(cfg, m2_year, m2_month)

    return render_template("index.html",
                           config=cfg,
                           month1=month1,
                           month2=month2,
                           today=today.isoformat())


@app.route("/api/config", methods=["GET"])
def api_get_config():
    return jsonify(load_config())


@app.route("/api/config", methods=["POST"])
def api_save_config():
    cfg = request.json
    save_config(cfg)
    return jsonify({"ok": True})


@app.route("/api/class", methods=["POST"])
def api_add_class():
    cfg = load_config()
    data = request.json
    cfg["classes"].append({
        "name": data.get("name", f"{len(cfg['classes'])+1}班"),
        "display": data.get("display", ""),
        "members": data.get("members", []),
    })
    save_config(cfg)
    return jsonify({"ok": True})


@app.route("/api/class/<int:idx>", methods=["PUT"])
def api_update_class(idx):
    cfg = load_config()
    data = request.json
    if 0 <= idx < len(cfg["classes"]):
        cfg["classes"][idx].update(data)
        save_config(cfg)
        return jsonify({"ok": True})
    return jsonify({"error": "index out of range"}), 400


@app.route("/api/class/<int:idx>", methods=["DELETE"])
def api_delete_class(idx):
    cfg = load_config()
    if 0 <= idx < len(cfg["classes"]):
        cfg["classes"].pop(idx)
        save_config(cfg)
        return jsonify({"ok": True})
    return jsonify({"error": "index out of range"}), 400


@app.route("/api/settings", methods=["POST"])
def api_update_settings():
    cfg = load_config()
    data = request.json
    for key in ("start_date", "admin", "year", "month"):
        if key in data:
            cfg[key] = data[key]
    save_config(cfg)
    return jsonify({"ok": True})


@app.route("/export/excel")
def export_excel():
    cfg = load_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "值班配置"
    ws.append(["班级名称", "显示人员", "值班人员"])
    for cls in cfg["classes"]:
        ws.append([
            cls["name"],
            cls["display"],
            "、".join(cls["members"]),
        ])
    # 列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="值班配置.xlsx")


@app.route("/import/excel", methods=["POST"])
def import_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "未上传文件"}), 400

    try:
        wb = load_workbook(file)
        ws = wb.active
        classes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[0] or "").strip()
            display = str(row[1] or "").strip()
            members_str = str(row[2] or "").strip()
            members = [m.strip() for m in members_str.split("、") if m.strip()]
            if name:
                classes.append({
                    "name": name,
                    "display": display,
                    "members": members,
                })
        cfg = load_config()
        cfg["classes"] = classes
        save_config(cfg)
        return jsonify({"ok": True, "count": len(classes)})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/month/<int:year>/<int:month>")
def api_month_data(year, month):
    cfg = load_config()
    return jsonify(build_month_data(cfg, year, month))


# ── 启动 ──────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
